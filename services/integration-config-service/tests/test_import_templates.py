"""Plantillas `.xlsx` descargables para cada uno de los 5 imports por Excel.

`GET /integrations/<recurso>/template` no toca base de datos ni exige autenticación —igual
que el resto de la fila de estos endpoints (`imports`, `siigo-syncs` sí la exigen, pero
descargar una hoja en blanco para llenar no expone nada de un tenant en particular)—, así que
se puede probar con un `TestClient` directo, sin mockear DB ni JWT.

Cada test valida tres cosas por endpoint: el `Content-Type`/`Content-Disposition` de una
descarga de archivo, y que la fila de encabezado del `.xlsx` generado coincide EXACTAMENTE con
las columnas que el `Import<X>UseCase` correspondiente espera (`REQUIRED_COLUMNS` +
`OPTIONAL_COLUMNS`) — para que la plantilla nunca se desincronice del parser real.
"""

from io import BytesIO

import pytest
from openpyxl import load_workbook

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def client(monkeypatch):
    monkeypatch.setenv("INTERNAL_SECRET", "secreto-de-prueba")
    from app.main import app

    from fastapi.testclient import TestClient

    return TestClient(app)


def _header_row(content: bytes) -> list[str]:
    workbook = load_workbook(filename=BytesIO(content))
    sheet = workbook.active
    return [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]


class TestTemplateEndpointsExist:
    """Cada plantilla responde 200 con un `.xlsx` descargable."""

    @pytest.mark.parametrize(
        "path,filename",
        [
            ("/api/v1/integrations/taxes/template", "plantilla-impuestos.xlsx"),
            ("/api/v1/integrations/cost-centers/template", "plantilla-centros-costo.xlsx"),
            ("/api/v1/integrations/payment-types/template", "plantilla-tipos-pago.xlsx"),
            ("/api/v1/integrations/products/template", "plantilla-productos.xlsx"),
            ("/api/v1/integrations/chart-accounts/template", "plantilla-plan-cuentas.xlsx"),
        ],
    )
    def test_descarga_xlsx_valido(self, client, path, filename):
        response = client.get(path)

        assert response.status_code == 200
        assert response.headers["content-type"] == XLSX_MEDIA_TYPE
        assert filename in response.headers["content-disposition"]
        # El cuerpo debe ser un .xlsx válido y no venir vacío.
        workbook = load_workbook(filename=BytesIO(response.content))
        assert workbook.active.max_row >= 1


class TestTemplateColumnsMatchParser:
    """Las columnas de la plantilla son exactamente las que el use case sabe leer."""

    def test_taxes_template_columns(self, client):
        from app.application.use_cases.import_taxes import ImportTaxesUseCase

        headers = set(_header_row(client.get("/api/v1/integrations/taxes/template").content))
        expected = ImportTaxesUseCase.REQUIRED_COLUMNS | ImportTaxesUseCase.OPTIONAL_COLUMNS
        assert headers == expected

    def test_cost_centers_template_columns(self, client):
        from app.application.use_cases.import_cost_centers import ImportCostCentersUseCase

        headers = set(
            _header_row(client.get("/api/v1/integrations/cost-centers/template").content)
        )
        expected = (
            ImportCostCentersUseCase.REQUIRED_COLUMNS | ImportCostCentersUseCase.OPTIONAL_COLUMNS
        )
        assert headers == expected

    def test_payment_types_template_columns(self, client):
        from app.application.use_cases.import_payment_types import ImportPaymentTypesUseCase

        headers = set(
            _header_row(client.get("/api/v1/integrations/payment-types/template").content)
        )
        expected = (
            ImportPaymentTypesUseCase.REQUIRED_COLUMNS
            | ImportPaymentTypesUseCase.OPTIONAL_COLUMNS
        )
        assert headers == expected

    def test_products_template_columns(self, client):
        from app.application.use_cases.import_products import ImportProductsUseCase

        headers = set(_header_row(client.get("/api/v1/integrations/products/template").content))
        expected = ImportProductsUseCase.REQUIRED_COLUMNS | ImportProductsUseCase.OPTIONAL_COLUMNS
        assert headers == expected

    def test_chart_accounts_template_columns(self, client):
        """El plan de cuentas es la excepcion a esta clase: su plantilla usa los encabezados
        reales de la exportacion de SIIGO (`Categoría`, `Relación con`...), no los nombres
        canonicos del use case — a proposito, para poder pegar ese archivo sin editarlo. Lo
        que debe seguir siendo cierto no es la igualdad literal sino que la plantilla se
        pueda volver a importar: sus encabezados, resueltos por alias, cubren las columnas
        obligatorias.
        """
        from app.application.use_cases.import_chart_accounts import ImportChartAccountsUseCase

        use_case = ImportChartAccountsUseCase
        headers = _header_row(client.get("/api/v1/integrations/chart-accounts/template").content)
        resolved = {
            use_case.HEADER_ALIASES.get(
                use_case._normalize_header(header), use_case._normalize_header(header)
            )
            for header in headers
        }
        assert use_case.REQUIRED_COLUMNS <= resolved

    @pytest.mark.parametrize(
        "path",
        [
            "/api/v1/integrations/taxes/template",
            "/api/v1/integrations/cost-centers/template",
            "/api/v1/integrations/payment-types/template",
            "/api/v1/integrations/products/template",
            "/api/v1/integrations/chart-accounts/template",
        ],
    )
    def test_template_has_no_prefilled_rows(self, client, path):
        """Una plantilla lista para llenar no trae datos: solo el encabezado.

        Antes la de impuestos venia con 5 filas de IVA/Impoconsumo precargadas — quien
        descargaba, adjuntaba y pulsaba Importar sin revisar terminaba creando esos 5
        impuestos sin haberlo pedido. Las otras cuatro ya llegaban en blanco; esta prueba
        fija esa misma garantia para las cinco.
        """
        workbook = load_workbook(filename=BytesIO(client.get(path).content))
        assert workbook.active.max_row == 1
