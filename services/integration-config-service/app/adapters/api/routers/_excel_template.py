"""Utilidades compartidas por los endpoints `GET .../template` de este servicio.

Vive en `adapters` (no en `domain`) porque es presentación pura: negrita, color de encabezado,
ancho de columnas, comentarios por celda y el `StreamingResponse` de descarga. Ningún caso de
uso depende de esto — cada router lo usa solo para construir el archivo `.xlsx`.

Mismo criterio de estilo que ya usa `xml-processor` para su plantilla de tarifas de retención
(`GET /catalog/retention-rates/template`): encabezado fijo y resaltado, una nota por columna, y
—cuando existe una tabla estándar razonable para precargar— filas de ejemplo; si no, la hoja de
datos llega limpia (solo encabezados) para no obligar a borrar filas de muestra antes de usarla.
"""

from io import BytesIO
from typing import Optional

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = "1F4E78"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def style_header(
    ws: Worksheet,
    notes: Optional[dict[str, str]] = None,
    widths: Optional[dict[str, int]] = None,
) -> None:
    """Encabezado en negrita con fondo de color, fila fija y una nota por columna."""
    notes = notes or {}
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.value in notes:
            cell.comment = Comment(notes[cell.value], "Abacus", height=140, width=320)
    for column, width in (widths or {}).items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"  # el encabezado sigue visible al desplazarse


def xlsx_response(workbook: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
