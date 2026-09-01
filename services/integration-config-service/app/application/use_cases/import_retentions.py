"""Importación de tarifas de ReteICA por municipio (Excel) — carga `integration_retentions`.

Repunta lo que antes vivía en `xml-processor` (`POST /catalog/retention-rates/imports?sheet=ica`
sobre `retention_ica_rates`) hacia el dueño de la tabla nueva. SIIGO no conoce municipios —su
ReteICA sincronizada es un porcentaje plano sin poder verificarse contra nada—, así que esta
es la ÚNICA vía por la que `integration_retentions` recibe filas `type='reteica'` de ahora en
adelante (el sync de SIIGO las descarta explícitamente, ver `sync_siigo_taxes.py`).

Mismas columnas y misma semántica de `replace` que ya tenía la importación de xml-processor,
para no obligar al contador a aprender un formato nuevo.
"""

import unicodedata
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook

from app.application.dto.retention import ImportRetentionsResponse
from app.domain.exceptions.base import ValidationException
from app.domain.services.ica_rate_units import UnidadesMezcladasError, verificar_unidad_coherente
from app.infrastructure.persistence.repositories.retention_repository import RetentionRepository

_ICA_SHEET = "ReteICA"

_ICA_HEADERS = {
    "codigo_municipio": "municipality_code",
    "codigo": "municipality_code",
    "municipality_code": "municipality_code",
    "municipio": "municipality_name",
    "nombre_municipio": "municipality_name",
    "municipality_name": "municipality_name",
    "tarifa": "percentage",
    "porcentaje": "percentage",
    "percentage": "percentage",
    "concepto": "retention_concept",
    "concepto_retencion": "retention_concept",
    "actividad": "retention_concept",
    "retention_concept": "retention_concept",
    "base_uvt": "minimum_base_uvt",
    "base_minima_uvt": "minimum_base_uvt",
    "minimum_base_uvt": "minimum_base_uvt",
}


class ImportRetentionsUseCase:
    def __init__(self, repository: RetentionRepository):
        self._repo = repository

    def execute(
        self, file_content: bytes, replace: bool = False, sheet_name: Optional[str] = None
    ) -> ImportRetentionsResponse:
        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 — openpyxl lanza tipos variados
            raise ValidationException("El archivo debe ser un Excel .xlsx válido") from exc

        sheets = {name.strip().lower(): name for name in workbook.sheetnames}
        nombre_hoja = (sheet_name or _ICA_SHEET).strip().lower()
        if nombre_hoja not in sheets:
            raise ValidationException(
                f"El archivo no trae la hoja '{sheet_name or _ICA_SHEET}' con las tarifas "
                "de ReteICA."
            )

        rows = self._parse(workbook[sheets[nombre_hoja]])
        if not rows:
            raise ValidationException(
                "El archivo no tiene filas para importar (la hoja está vacía)."
            )

        try:
            verificar_unidad_coherente(rows)
        except UnidadesMezcladasError as exc:
            raise ValidationException(str(exc)) from exc

        ica_loaded = self._repo.upsert_ica_rows(rows, replace=replace)
        return ImportRetentionsResponse(
            ica_loaded=ica_loaded, retentions=self._repo.list(type="reteica")
        )

    # ── Parseo ─────────────────────────────────────────────────────────────────
    def _parse(self, sheet) -> list[dict[str, Any]]:
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise ValidationException(f"La hoja '{sheet.title}' está vacía") from exc

        header_map: dict[str, int] = {}
        for idx, value in enumerate(header):
            key = self._normalize(value)
            field = _ICA_HEADERS.get(key)
            if field and field not in header_map:
                header_map[field] = idx

        required = {"municipality_code", "percentage"}
        missing = required - set(header_map)
        if missing:
            raise ValidationException(
                f"La hoja '{sheet.title}' no tiene las columnas requeridas: "
                f"{', '.join(sorted(missing))}"
            )

        rows: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()
        for line, raw in enumerate(rows_iter, start=2):
            if raw is None or all(c is None or str(c).strip() == "" for c in raw):
                continue
            code = self._text(raw, header_map.get("municipality_code"))
            if not code or str(code).startswith("#"):
                continue
            concepto = (self._text(raw, header_map.get("retention_concept")) or "todos").strip().lower()
            clave = (code, concepto)
            if clave in seen:
                raise ValidationException(
                    f"Fila {line}: ya hay una tarifa para el municipio '{code}' y el concepto "
                    f"'{concepto}'. Un municipio puede tener varias tarifas, pero solo una por "
                    "concepto."
                )
            seen.add(clave)
            rows.append(
                {
                    "municipality_code": code,
                    "municipality_name": self._text(raw, header_map.get("municipality_name")),
                    "retention_concept": concepto,
                    "percentage": self._num(raw, header_map.get("percentage"), line, "tarifa"),
                    "minimum_base_uvt": self._num(
                        raw, header_map.get("minimum_base_uvt"), line, "base_uvt", required=False
                    ),
                }
            )
        return rows

    @staticmethod
    def _normalize(value: Any) -> str:
        if value is None:
            return ""
        text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
        return text.strip().lower().replace(" ", "_")

    @staticmethod
    def _text(row, index: Optional[int]) -> Optional[str]:
        if index is None or index >= len(row) or row[index] is None:
            return None
        valor = row[index]
        if isinstance(valor, float) and valor.is_integer():
            valor = int(valor)
        text = str(valor).strip()
        return text or None

    @staticmethod
    def _num(
        row, index: Optional[int], line: int, field: str, *, required: bool = True
    ) -> Optional[float]:
        if index is None or index >= len(row) or row[index] is None or str(row[index]).strip() == "":
            if required:
                raise ValidationException(f"Fila {line}: la columna '{field}' es obligatoria")
            return None
        raw = str(row[index]).strip().replace("%", "").replace(",", ".")
        try:
            return float(raw)
        except ValueError as exc:
            raise ValidationException(
                f"Fila {line}: '{field}' debe ser numérico, se recibió '{row[index]}'"
            ) from exc
