from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook

from app.application.dto.tax import ImportTaxesResponse
from app.domain.exceptions.base import ValidationException
from app.infrastructure.persistence.repositories.tax_repository import TaxRepository


class ImportTaxesUseCase:
    REQUIRED_COLUMNS = {"name", "type", "percentage"}
    OPTIONAL_COLUMNS = {"active"}

    def __init__(self, repository: TaxRepository):
        self.repository = repository

    def execute(
        self,
        file_content: bytes,
        sheet_name: Optional[str] = None,
    ) -> ImportTaxesResponse:
        taxes = self._parse_excel(file_content=file_content, sheet_name=sheet_name)
        imported = self.repository.upsert_many(taxes)
        return ImportTaxesResponse(imported=imported, taxes=self.repository.list())

    def _parse_excel(self, file_content: bytes, sheet_name: Optional[str]) -> list[dict[str, Any]]:
        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationException("The uploaded file must be a valid .xlsx Excel file") from exc

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValidationException(f"Sheet not found: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ValidationException("The Excel file is empty") from exc

        headers = [self._normalize_header(value) for value in header_row]
        header_map = {name: index for index, name in enumerate(headers) if name}
        missing = self.REQUIRED_COLUMNS - set(header_map)
        if missing:
            raise ValidationException(f"Missing required columns: {', '.join(sorted(missing))}")

        taxes: list[dict[str, Any]] = []
        seen_names: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            if self._is_empty_row(row):
                continue

            name = self._cell(row, header_map["name"])
            type_ = self._cell(row, header_map["type"])
            percentage = self._cell(row, header_map["percentage"])

            if name is None or str(name).strip() == "":
                raise ValidationException(f"Row {row_number}: name is required")
            if type_ is None or str(type_).strip() == "":
                raise ValidationException(f"Row {row_number}: type is required")
            if percentage is None or str(percentage).strip() == "":
                raise ValidationException(f"Row {row_number}: percentage is required")

            try:
                percentage_value = float(str(percentage).strip())
            except ValueError as exc:
                raise ValidationException(
                    f"Row {row_number}: percentage must be a number"
                ) from exc

            normalized_name = str(name).strip()
            if normalized_name in seen_names:
                raise ValidationException(f"Row {row_number}: duplicated name '{normalized_name}'")
            seen_names.add(normalized_name)

            taxes.append(
                {
                    "name": normalized_name,
                    "type": str(type_).strip(),
                    "percentage": percentage_value,
                    "active": self._as_bool(
                        self._optional_cell(row, header_map, "active"), default=True
                    ),
                }
            )

        if not taxes:
            raise ValidationException("The Excel file does not contain tax rows")
        return taxes

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return str(value).strip().lower() if value is not None else ""

    @staticmethod
    def _is_empty_row(row: tuple) -> bool:
        return all(value is None or str(value).strip() == "" for value in row)

    @staticmethod
    def _cell(row: tuple, index: int) -> Any:
        return row[index] if index < len(row) else None

    def _optional_cell(self, row: tuple, header_map: dict[str, int], column: str) -> Any:
        index = header_map.get(column)
        return self._cell(row, index) if index is not None else None

    def _as_bool(self, value: Any, default: bool) -> bool:
        if value is None or str(value).strip() == "":
            return default
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        if normalized in {"true", "1", "yes", "y", "si", "sí", "x"}:
            return True
        if normalized in {"false", "0", "no", "n"}:
            return False
        raise ValidationException("active must be true/false")
