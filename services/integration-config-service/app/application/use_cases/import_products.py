from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook

from app.application.dto.product import ImportProductsResponse
from app.domain.exceptions.base import ValidationException
from app.infrastructure.persistence.repositories.product_repository import ProductRepository

VALID_TYPES = {"product", "service"}
# El valor se guarda en ingles (asi lo espera el resto del dominio), pero se escribe en
# espanol en la plantilla: nadie que llena un Excel en espanol escribe "product"/"service".
TYPE_ALIASES = {"producto": "product", "servicio": "service"}


class ImportProductsUseCase:
    REQUIRED_COLUMNS = {"código", "tipo", "descripción"}
    OPTIONAL_COLUMNS = {"activo"}
    # `code`/`type`/`description`/`active` siguen aceptandose: son el encabezado con que
    # este endpoint funciono antes de que la plantilla pasara a espanol, y romperian un
    # archivo ya armado con ellos. El encabezado explicito en espanol siempre gana sobre el
    # alias, y se acepta tanto con tilde como sin ella.
    HEADER_ALIASES = {
        "code": "código",
        "codigo": "código",
        "type": "tipo",
        "description": "descripción",
        "descripcion": "descripción",
        "active": "activo",
    }

    def __init__(self, repository: ProductRepository):
        self.repository = repository

    def execute(
        self,
        file_content: bytes,
        sheet_name: Optional[str] = None,
        mode: str = "upsert",
    ) -> ImportProductsResponse:
        if mode not in ("upsert", "replace"):
            raise ValidationException("mode must be 'upsert' or 'replace'")
        products = self._parse_excel(file_content=file_content, sheet_name=sheet_name)
        imported = self.repository.upsert_many(products=products, replace=(mode == "replace"))
        return ImportProductsResponse(
            imported=imported,
            products=self.repository.list(),
        )

    def _parse_excel(self, file_content: bytes, sheet_name: Optional[str]) -> list[dict[str, Any]]:
        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationException("The uploaded file must be a valid .xlsx Excel file") from exc

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValidationException(f"Sheet not found: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ValidationException("The Excel file is empty") from exc

        headers = [self._normalize_header(value) for value in header_row]
        header_map = {name: index for index, name in enumerate(headers) if name}
        # Un encabezado en espanol explicito siempre gana sobre su alias en ingles.
        for index, name in enumerate(headers):
            canonical = self.HEADER_ALIASES.get(name)
            if canonical and canonical not in header_map:
                header_map[canonical] = index
        missing = self.REQUIRED_COLUMNS - set(header_map)
        if missing:
            raise ValidationException(f"Missing required columns: {', '.join(sorted(missing))}")

        products: list[dict[str, Any]] = []
        seen_codes = set()
        for row_number, row in enumerate(rows, start=2):
            if self._is_empty_row(row):
                continue

            code = self._cell(row, header_map["código"])
            type_val = self._cell(row, header_map["tipo"])
            description = self._cell(row, header_map["descripción"])

            if code is None or str(code).strip() == "":
                raise ValidationException(f"Row {row_number}: code is required")
            if type_val is None or str(type_val).strip() == "":
                raise ValidationException(f"Row {row_number}: type is required")
            if description is None or str(description).strip() == "":
                raise ValidationException(f"Row {row_number}: description is required")

            normalized_code = self._as_code(code)
            if normalized_code in seen_codes:
                raise ValidationException(f"Row {row_number}: duplicated code {normalized_code}")
            seen_codes.add(normalized_code)

            normalized_type = str(type_val).strip().lower()
            normalized_type = TYPE_ALIASES.get(normalized_type, normalized_type)
            if normalized_type not in VALID_TYPES:
                raise ValidationException(
                    f"Row {row_number}: type must be 'producto'/'product' or "
                    f"'servicio'/'service', got '{normalized_type}'"
                )

            raw_payload = {
                column: self._cell(row, index)
                for column, index in header_map.items()
                if column in self.REQUIRED_COLUMNS or column in self.OPTIONAL_COLUMNS
            }
            products.append(
                {
                    "code": normalized_code,
                    "type": normalized_type,
                    "description": str(description).strip(),
                    "active": self._as_bool(
                        self._optional_cell(row, header_map, "activo"), default=True
                    ),
                    "raw_payload": raw_payload,
                }
            )

        if not products:
            raise ValidationException("The Excel file does not contain product rows")
        return products

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return str(value).strip().lower() if value is not None else ""

    @staticmethod
    def _is_empty_row(row: tuple) -> bool:
        return all(value is None or str(value).strip() == "" for value in row)

    @staticmethod
    def _cell(row: tuple, index: int) -> Any:
        return row[index] if index < len(row) else None

    def _optional_cell(self, row: tuple, header_map: dict[str, int], column: str) -> Any:
        index = header_map.get(column)
        return self._cell(row, index) if index is not None else None

    @staticmethod
    def _as_code(value: Any) -> str:
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _as_bool(self, value: Any, default: bool) -> bool:
        if value is None or str(value).strip() == "":
            return default
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        if normalized in {"true", "1", "yes", "y", "si", "sí", "x"}:
            return True
        if normalized in {"false", "0", "no", "n"}:
            return False
        raise ValidationException("active must be true/false")
