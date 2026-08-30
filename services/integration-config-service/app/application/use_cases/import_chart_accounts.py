from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook

from app.application.dto.chart_account import ImportChartAccountsResponse
from app.domain.exceptions.base import ValidationException
from app.infrastructure.clients.xml_processor_client import XmlProcessorClient
from app.infrastructure.persistence.repositories.chart_account_repository import (
    ChartAccountRepository,
)


class ImportChartAccountsUseCase:
    REQUIRED_COLUMNS = {"code", "name"}
    OPTIONAL_COLUMNS = {
        "external_id",
        "account_type",
        "level",
        "parent_code",
        "accepts_movements",
        "active",
    }
    # Las exportaciones reales del plan de cuentas de SIIGO traen los encabezados en
    # español. Se mapean a los nombres canónicos para no obligar al contador a renombrar
    # columnas antes de subir el archivo. Solo aliases inequívocos: p. ej. «nivel
    # agrupación» NO es «nivel», por eso la coincidencia es exacta y no por prefijo.
    HEADER_ALIASES = {
        "código": "code",
        "codigo": "code",
        "nombre": "name",
        "tipo de cuenta": "account_type",
        "nivel": "level",
        "cuenta padre": "parent_code",
        "código padre": "parent_code",
        "codigo padre": "parent_code",
        "activo": "active",
        "activa": "active",
    }
    # Longitud del código PUC → nivel jerárquico (clase, grupo, cuenta, subcuenta,
    # auxiliar…). Se usa solo cuando el archivo no trae una columna `level` explícita.
    _LEVEL_BY_CODE_LENGTH = {1: 1, 2: 2, 4: 3, 6: 4, 8: 5, 10: 6}

    def __init__(
        self,
        repository: ChartAccountRepository,
        tenant_slug: str,
        xml_processor_client: Optional[XmlProcessorClient] = None,
    ):
        self.repository = repository
        self.tenant_slug = tenant_slug
        self.xml_processor_client = xml_processor_client or XmlProcessorClient()

    def execute(
        self,
        file_content: bytes,
        sheet_name: Optional[str] = None,
        mode: str = "upsert",
    ) -> ImportChartAccountsResponse:
        if mode not in ("upsert", "replace"):
            raise ValidationException("mode must be 'upsert' or 'replace'")
        accounts = self._parse_excel(file_content=file_content, sheet_name=sheet_name)
        if mode == "replace":
            self.repository.delete_all()
        imported = self.repository.upsert_many(accounts=accounts)
        movements = self._recalculate_accepts_movements(accounts)

        # El catálogo que consumen el selector de cuenta y el llm-service vive en
        # `puc_accounts` (xml-processor). Sin esta proyección la importación quedaría
        # aislada en `integration_chart_accounts` y el selector seguiría vacío.
        #
        # `accepts_movements` viaja con el resto: es lo que permite a xml-processor rechazar
        # una imputación contra una cuenta agrupadora. Antes se calculaba aquí y se quedaba
        # aquí, así que ni el selector ni la validación del LLM podían distinguir una
        # auxiliar de un grupo.
        self.xml_processor_client.project_puc_accounts(
            tenant_slug=self.tenant_slug,
            items=[
                {
                    "code": a["code"],
                    "name": a["name"],
                    "level": a.get("level"),
                    "active": a.get("active", True),
                    "accepts_movements": (
                        a["accepts_movements"]
                        if a.get("accepts_movements") is not None
                        else movements.get(a["code"])
                    ),
                }
                for a in accounts
            ],
        )
        return ImportChartAccountsResponse(
            imported=imported,
            accounts=self.repository.list(),
        )

    def _recalculate_accepts_movements(self, accounts: list[dict[str, Any]]) -> dict[str, bool]:
        """Marca como imputable cada cuenta hoja y retorna el mapa calculado.

        Se devuelve —además de persistirse— porque la proyección hacia `puc_accounts`
        necesita el mismo dato: recalcularlo allí duplicaría el criterio y abriría la
        puerta a que ambos catálogos discrepen.
        """
        codes = [a["code"] for a in accounts]
        numeric_codes = {c for c in codes if c.isdigit() and len(c) >= 4}
        # A code is a leaf if no other code in the dataset starts with it and is longer.
        # This handles branches with different max depths (e.g. 22xxxx tops at 8 digits,
        # while 51xxxx tops at 10 digits — both sets of leaves get accepts_movements=True).
        movements = {
            c: not any(other.startswith(c) and len(other) > len(c) for other in numeric_codes)
            for c in codes
            if c.isdigit() and len(c) >= 4
        }
        # Non-numeric codes (e.g. alphanumeric) never accept movements.
        for c in codes:
            if not (c.isdigit() and len(c) >= 4):
                movements[c] = False
        self.repository.set_accepts_movements(movements)
        return movements

    def _parse_excel(self, file_content: bytes, sheet_name: Optional[str]) -> list[dict[str, Any]]:
        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationException("The uploaded file must be a valid .xlsx Excel file") from exc

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValidationException(f"Sheet not found: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ValidationException("The Excel file is empty") from exc

        headers = [self._normalize_header(value) for value in header_row]
        header_map = {name: index for index, name in enumerate(headers) if name}
        # Un encabezado canónico explícito siempre gana sobre su alias en español.
        for index, name in enumerate(headers):
            canonical = self.HEADER_ALIASES.get(name)
            if canonical and canonical not in header_map:
                header_map[canonical] = index
        missing = self.REQUIRED_COLUMNS - set(header_map)
        if missing:
            raise ValidationException(f"Missing required columns: {', '.join(sorted(missing))}")

        accounts: list[dict[str, Any]] = []
        seen_codes = set()
        for row_number, row in enumerate(rows, start=2):
            if self._is_empty_row(row):
                continue

            code = self._cell(row, header_map["code"])
            name = self._cell(row, header_map["name"])
            if code is None or str(code).strip() == "":
                raise ValidationException(f"Row {row_number}: code is required")
            if name is None or str(name).strip() == "":
                raise ValidationException(f"Row {row_number}: name is required")

            normalized_code = self._as_code(code)
            if normalized_code in seen_codes:
                raise ValidationException(f"Row {row_number}: duplicated code {normalized_code}")
            seen_codes.add(normalized_code)

            explicit_level = self._as_optional_int(
                self._optional_cell(row, header_map, "level"), row_number, "level"
            )

            raw_payload = {
                column: self._cell(row, index)
                for column, index in header_map.items()
                if column in self.REQUIRED_COLUMNS or column in self.OPTIONAL_COLUMNS
            }
            accounts.append(
                {
                    "external_id": self._as_optional_text(
                        self._optional_cell(row, header_map, "external_id")
                    ),
                    "code": normalized_code,
                    "name": str(name).strip(),
                    "account_type": self._as_optional_text(
                        self._optional_cell(row, header_map, "account_type")
                    ),
                    "level": explicit_level
                    if explicit_level is not None
                    else self._derive_level(normalized_code),
                    "parent_code": self._as_optional_code(
                        self._optional_cell(row, header_map, "parent_code")
                    ),
                    "accepts_movements": self._as_optional_bool(
                        self._optional_cell(row, header_map, "accepts_movements"),
                        row_number,
                        "accepts_movements",
                    ),
                    "active": self._as_bool(
                        self._optional_cell(row, header_map, "active"), default=True
                    ),
                    "raw_payload": raw_payload,
                }
            )

        if not accounts:
            raise ValidationException("The Excel file does not contain account rows")
        return accounts

    @classmethod
    def _derive_level(cls, code: str) -> Optional[int]:
        """Nivel jerárquico inferido de la longitud del código PUC.

        Solo aplica cuando el archivo no trae la columna `level`. Devuelve None para
        códigos no numéricos o de longitud inesperada, para no inventar jerarquías.
        """
        if not code.isdigit():
            return None
        return cls._LEVEL_BY_CODE_LENGTH.get(len(code))

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return str(value).strip().lower() if value is not None else ""

    @staticmethod
    def _is_empty_row(row: tuple) -> bool:
        return all(value is None or str(value).strip() == "" for value in row)

    @staticmethod
    def _cell(row: tuple, index: int) -> Any:
        return row[index] if index < len(row) else None

    def _optional_cell(self, row: tuple, header_map: dict[str, int], column: str) -> Any:
        index = header_map.get(column)
        return self._cell(row, index) if index is not None else None

    @staticmethod
    def _as_code(value: Any) -> str:
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _as_optional_code(self, value: Any) -> Optional[str]:
        if value is None or str(value).strip() == "":
            return None
        return self._as_code(value)

    @staticmethod
    def _as_optional_text(value: Any) -> Optional[str]:
        if value is None or str(value).strip() == "":
            return None
        return str(value).strip()

    @staticmethod
    def _as_optional_int(value: Any, row_number: int, column: str) -> Optional[int]:
        if value is None or str(value).strip() == "":
            return None
        try:
            return int(value)
        except (TypeError, ValueError) as exc:
            raise ValidationException(f"Row {row_number}: {column} must be an integer") from exc

    def _as_optional_bool(self, value: Any, row_number: int, column: str) -> Optional[bool]:
        if value is None or str(value).strip() == "":
            return None
        try:
            return self._parse_bool(value)
        except ValueError as exc:
            raise ValidationException(f"Row {row_number}: {column} must be true/false") from exc

    def _as_bool(self, value: Any, default: bool) -> bool:
        if value is None or str(value).strip() == "":
            return default
        return self._parse_bool(value)

    @staticmethod
    def _parse_bool(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        if normalized in {"true", "1", "yes", "y", "si", "sí", "x"}:
            return True
        if normalized in {"false", "0", "no", "n"}:
            return False
        raise ValueError("invalid boolean")
