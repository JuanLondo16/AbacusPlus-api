from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from app.application.dto.cost_center import ImportCostCentersResponse
from app.domain.exceptions.base import ValidationException
from app.infrastructure.persistence.repositories.cost_center_repository import CostCenterRepository


class ImportCostCentersUseCase:
    REQUIRED_COLUMNS = {"code", "name"}
    OPTIONAL_COLUMNS = {"external_id", "active"}

    def __init__(self, repository: CostCenterRepository):
        self.repository = repository

    def execute(
        self,
        provider: str,
        account_key: str,
        file_content: bytes,
        sheet_name: Optional[str] = None,
    ) -> ImportCostCentersResponse:
        provider = provider.strip().lower()
        account_key = account_key.strip()
        if not provider:
            raise ValidationException("provider is required")
        if not account_key:
            raise ValidationException("account_key is required")

        cost_centers = self._parse_excel(file_content=file_content, sheet_name=sheet_name)
        imported = self.repository.upsert_many(provider=provider, account_key=account_key, cost_centers=cost_centers)
        return ImportCostCentersResponse(
            provider=provider,
            account_key=account_key,
            imported=imported,
            cost_centers=self.repository.list(provider=provider, account_key=account_key),
        )

    def _parse_excel(self, file_content: bytes, sheet_name: Optional[str]) -> List[Dict[str, Any]]:
        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationException("The uploaded file must be a valid .xlsx Excel file") from exc

        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValidationException(f"Sheet not found: {sheet_name}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ValidationException("The Excel file is empty") from exc

        headers = [self._normalize_header(value) for value in header_row]
        header_map = {name: index for index, name in enumerate(headers) if name}
        missing = self.REQUIRED_COLUMNS - set(header_map)
        if missing:
            raise ValidationException(f"Missing required columns: {', '.join(sorted(missing))}")

        cost_centers: List[Dict[str, Any]] = []
        seen_codes = set()
        for row_number, row in enumerate(rows, start=2):
            if self._is_empty_row(row):
                continue

            code = self._cell(row, header_map["code"])
            name = self._cell(row, header_map["name"])
            if code is None or str(code).strip() == "":
                raise ValidationException(f"Row {row_number}: code is required")
            if name is None or str(name).strip() == "":
                raise ValidationException(f"Row {row_number}: name is required")

            normalized_code = self._as_code(code)
            if normalized_code in seen_codes:
                raise ValidationException(f"Row {row_number}: duplicated code {normalized_code}")
            seen_codes.add(normalized_code)

            raw_payload = {
                column: self._cell(row, index)
                for column, index in header_map.items()
                if column in self.REQUIRED_COLUMNS or column in self.OPTIONAL_COLUMNS
            }
            cost_centers.append(
                {
                    "external_id": self._as_optional_text(self._optional_cell(row, header_map, "external_id")),
                    "code": normalized_code,
                    "name": str(name).strip(),
                    "active": self._as_bool(self._optional_cell(row, header_map, "active"), default=True),
                    "raw_payload": raw_payload,
                }
            )

        if not cost_centers:
            raise ValidationException("The Excel file does not contain cost center rows")
        return cost_centers

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return str(value).strip().lower() if value is not None else ""

    @staticmethod
    def _is_empty_row(row: tuple) -> bool:
        return all(value is None or str(value).strip() == "" for value in row)

    @staticmethod
    def _cell(row: tuple, index: int) -> Any:
        return row[index] if index < len(row) else None

    def _optional_cell(self, row: tuple, header_map: Dict[str, int], column: str) -> Any:
        index = header_map.get(column)
        return self._cell(row, index) if index is not None else None

    @staticmethod
    def _as_code(value: Any) -> str:
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _as_optional_text(value: Any) -> Optional[str]:
        if value is None or str(value).strip() == "":
            return None
        return str(value).strip()

    def _as_bool(self, value: Any, default: bool) -> bool:
        if value is None or str(value).strip() == "":
            return default
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        if normalized in {"true", "1", "yes", "y", "si", "sí", "x"}:
            return True
        if normalized in {"false", "0", "no", "n"}:
            return False
        raise ValidationException("active must be true/false")
