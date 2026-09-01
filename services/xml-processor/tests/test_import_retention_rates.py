"""Importación de tarifas de retención desde Excel (RF-08).

Se prueba el parseo y la validación del caso de uso con un repositorio falso —sin base de
datos ni red—. El archivo `.xlsx` se construye en memoria con openpyxl.
"""

from io import BytesIO

import pytest
from app.application.use_cases.import_retention_rates import ImportRetentionRatesUseCase
from app.domain.exceptions.base import ValidationException
from openpyxl import Workbook


class _FakeRepo:
    def __init__(self):
        self.fuente = None
        self.ica = None
        #: Modo con el que el caso de uso invocó al repositorio. None = no se llamó.
        self.replace = None

    def import_rates(self, fuente_rows, ica_rows, *, replace=False):
        self.fuente = fuente_rows
        self.ica = ica_rows
        self.replace = replace
        f = len(fuente_rows) if fuente_rows is not None else 0
        i = len(ica_rows) if ica_rows is not None else 0
        return f, i


def _xlsx(fuente=None, ica=None, ica_headers=None):
    """Crea un .xlsx en memoria. `fuente`/`ica` son listas de filas (sin encabezado).

    `ica_headers` permite probar la hoja CON columna de concepto, que es como la carga el
    contador cuando su municipio tiene una tarifa por actividad.
    """
    wb = Workbook()
    wb.remove(wb.active)  # quita la hoja por defecto
    if fuente is not None:
        ws = wb.create_sheet("ReteFuente")
        ws.append(["concepto", "tipo_contribuyente", "base_uvt", "base_pesos", "tarifa"])
        for row in fuente:
            ws.append(row)
    if ica is not None:
        ws = wb.create_sheet("ReteICA")
        ws.append(ica_headers or ["codigo_municipio", "municipio", "tarifa"])
        for row in ica:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_importa_ambas_hojas():
    repo = _FakeRepo()
    content = _xlsx(
        fuente=[["Compras generales", "declarante", 10, 523740, 2.5]],
        ica=[["11001", "Bogotá D.C.", 0.966]],
    )
    result = ImportRetentionRatesUseCase(repo).execute(content)
    assert result == {"fuente_loaded": 1, "ica_loaded": 1}
    assert repo.fuente[0]["retention_concept"] == "Compras generales"
    assert repo.fuente[0]["taxpayer_type"] == "declarante"
    assert repo.fuente[0]["rate_percentage"] == 2.5
    assert repo.ica[0] == {
        "municipality_code": "11001",
        "municipality_name": "Bogotá D.C.",
        "retention_concept": "todos",
        "percentage": 0.966,
        "minimum_base_uvt": None,
    }


def test_hoja_ica_vacia_se_omite_y_fuente_se_importa():
    # Escenario plantilla: ReteFuente llena, ReteICA presente pero sin filas → no toca ReteICA.
    repo = _FakeRepo()
    content = _xlsx(
        fuente=[["Compras generales", "declarante", 10, 523740, 2.5]],
        ica=[],  # hoja presente, sin datos
    )
    result = ImportRetentionRatesUseCase(repo).execute(content)
    assert result == {"fuente_loaded": 1, "ica_loaded": 0}
    assert repo.ica is None  # no se vació la tabla de ReteICA


def test_solo_ica_no_toca_fuente():
    repo = _FakeRepo()
    content = _xlsx(ica=[["11001", "Bogotá D.C.", 0.966]])
    result = ImportRetentionRatesUseCase(repo).execute(content)
    assert result == {"fuente_loaded": 0, "ica_loaded": 1}
    assert repo.fuente is None  # sin hoja ReteFuente → no se toca esa tabla


def test_tarifa_con_simbolo_y_coma_decimal():
    repo = _FakeRepo()
    content = _xlsx(ica=[["11001", "Bogotá", "0,966%"]])
    ImportRetentionRatesUseCase(repo).execute(content)
    assert repo.ica[0]["percentage"] == 0.966


def test_tipo_contribuyente_invalido_falla():
    repo = _FakeRepo()
    content = _xlsx(fuente=[["Compras", "persona_natural_x", 10, 523740, 2.5]])
    with pytest.raises(ValidationException) as exc:
        ImportRetentionRatesUseCase(repo).execute(content)
    assert "tipo_contribuyente" in str(exc.value).lower()


def test_municipio_duplicado_falla():
    repo = _FakeRepo()
    content = _xlsx(ica=[["11001", "Bogotá", 0.966], ["11001", "Bogotá otra", 1.0]])
    with pytest.raises(ValidationException):
        ImportRetentionRatesUseCase(repo).execute(content)


def test_columna_obligatoria_faltante_falla():
    # ReteICA sin columna de tarifa
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("ReteICA")
    ws.append(["codigo_municipio", "municipio"])
    ws.append(["11001", "Bogotá"])
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(ValidationException) as exc:
        ImportRetentionRatesUseCase(_FakeRepo()).execute(buf.getvalue())
    assert "percentage" in str(exc.value).lower() or "tarifa" in str(exc.value).lower()


def test_sin_hojas_reconocidas_falla():
    wb = Workbook()  # solo la hoja por defecto "Sheet"
    buf = BytesIO()
    wb.save(buf)
    with pytest.raises(ValidationException):
        ImportRetentionRatesUseCase(_FakeRepo()).execute(buf.getvalue())


def test_archivo_no_xlsx_falla():
    with pytest.raises(ValidationException):
        ImportRetentionRatesUseCase(_FakeRepo()).execute(b"esto no es un excel")


# ── Modo de importación: upsert vs reemplazo ───────────────────────────────────


def test_por_defecto_no_reemplaza():
    """El modo seguro es el implícito: subir dos conceptos no puede vaciar la tabla.

    Antes ReteFuente reemplazaba siempre, así que una hoja con dos filas borraba la tabla
    nacional entera sin avisar. El valor por defecto invierte ese riesgo.
    """
    repo = _FakeRepo()
    content = _xlsx(fuente=[["Compras generales", "declarante", 10, 523740, 2.5]])

    ImportRetentionRatesUseCase(repo).execute(content)

    assert repo.replace is False


def test_el_modo_reemplazo_se_propaga_al_repositorio():
    repo = _FakeRepo()
    content = _xlsx(fuente=[["Compras generales", "declarante", 10, 523740, 2.5]])

    ImportRetentionRatesUseCase(repo).execute(content, replace=True)

    assert repo.replace is True


def test_reemplazar_no_afecta_a_la_hoja_que_no_viene():
    """Con `replace=True` y solo ReteICA, la tabla nacional debe quedar intacta.

    Lo garantiza el contrato con el repositorio: `fuente_rows=None` significa «no tocar», y
    el modo de reemplazo no puede cambiar eso. Sin esta regla, subir un municipio borraría
    todas las tarifas de ReteFuente.
    """
    repo = _FakeRepo()
    content = _xlsx(ica=[["11001", "Bogotá D.C.", 0.966]])

    ImportRetentionRatesUseCase(repo).execute(content, replace=True)

    assert repo.fuente is None
    assert repo.ica is not None
    assert repo.replace is True


# ── Duplicados dentro de la hoja ───────────────────────────────────────────────


def test_rechaza_conceptos_repetidos_en_retefuente():
    """(concepto, tipo_contribuyente) es la clave única: repetirla es un error del usuario.

    En reemplazo la inserción fallaría contra la restricción única con un mensaje ilegible;
    en upsert la segunda fila pisaría a la primera en silencio.
    """
    repo = _FakeRepo()
    content = _xlsx(
        fuente=[
            ["Compras generales", "declarante", 10, 523740, 2.5],
            ["Compras generales", "declarante", 10, 523740, 3.5],
        ]
    )

    with pytest.raises(ValidationException):
        ImportRetentionRatesUseCase(repo).execute(content)


def test_el_mismo_concepto_con_distinto_tipo_si_es_valido():
    """Un concepto puede tener tarifa distinta según el tipo de contribuyente."""
    repo = _FakeRepo()
    content = _xlsx(
        fuente=[
            ["Compras generales", "declarante", 10, 523740, 2.5],
            ["Compras generales", "no_declarante", 10, 523740, 3.5],
        ]
    )

    resultado = ImportRetentionRatesUseCase(repo).execute(content)

    assert resultado["fuente_loaded"] == 2


# ── RF-08 · ReteICA con varias tarifas por municipio ──────────────────────────
#
# La tarifa de ReteICA la fija el CONCEPTO de la operación, no el municipio: en Bogotá hay una
# banda por actividad. Hasta ahora la tabla admitía una sola tarifa por municipio, así que era
# imposible cargarlas y el sistema no podía elegir cuál aplicar.

_ICA_CON_CONCEPTO = ["codigo_municipio", "municipio", "concepto", "tarifa"]


def test_un_municipio_admite_varias_tarifas_una_por_concepto():
    repo = _FakeRepo()
    content = _xlsx(
        ica=[
            ["11001", "Bogotá D.C.", "servicios", 0.966],
            ["11001", "Bogotá D.C.", "compras", 1.104],
            ["11001", "Bogotá D.C.", "honorarios", 0.966],
        ],
        ica_headers=_ICA_CON_CONCEPTO,
    )

    result = ImportRetentionRatesUseCase(repo).execute(content)

    assert result["ica_loaded"] == 3
    assert [r["retention_concept"] for r in repo.ica] == ["servicios", "compras", "honorarios"]
    assert {r["municipality_code"] for r in repo.ica} == {"11001"}


def test_el_concepto_se_normaliza_a_minusculas():
    """El contador escribe «Servicios» o «SERVICIOS»; deben ser el mismo concepto.

    Sin normalizar, dos grafías del mismo concepto conviven como filas distintas y el
    emparejamiento contra la descripción de la factura se vuelve azaroso.
    """
    repo = _FakeRepo()
    content = _xlsx(
        ica=[["11001", "Bogotá D.C.", "  Servicios  ", 0.966]],
        ica_headers=_ICA_CON_CONCEPTO,
    )

    ImportRetentionRatesUseCase(repo).execute(content)

    assert repo.ica[0]["retention_concept"] == "servicios"


def test_se_rechaza_el_mismo_municipio_y_concepto_dos_veces():
    """Repetir el municipio es normal; repetir (municipio, concepto) es ambiguo.

    Dos tarifas para el mismo caso no dicen cuál vale, y elegir una en silencio sería
    inventar una decisión tributaria.
    """
    repo = _FakeRepo()
    content = _xlsx(
        ica=[
            ["11001", "Bogotá D.C.", "servicios", 0.966],
            ["11001", "Bogotá D.C.", "servicios", 1.104],
        ],
        ica_headers=_ICA_CON_CONCEPTO,
    )

    with pytest.raises(ValidationException) as exc:
        ImportRetentionRatesUseCase(repo).execute(content)

    assert "concepto" in str(exc.value).lower()


def test_una_plantilla_antigua_sin_concepto_sigue_siendo_valida():
    """Compatibilidad: la columna es opcional y su ausencia significa «aplica a todo»."""
    repo = _FakeRepo()
    content = _xlsx(ica=[["11001", "Bogotá D.C.", 0.966], ["05001", "Medellín", 0.7]])

    result = ImportRetentionRatesUseCase(repo).execute(content)

    assert result["ica_loaded"] == 2
    assert {r["retention_concept"] for r in repo.ica} == {"todos"}


# ── Base mínima por municipio · el ICA es territorial ─────────────────────────
#
# Cada municipio fija su propio tope y no hay uniformidad nacional: Bogotá pide 4 UVT en
# servicios y 27 en compras; Cali 3 y 15; Bucaramanga 25 y 50. Con el valor de Bogotá fijo en
# el código, contabilizar en Bucaramanga proponía ReteICA sobre facturas que no la causan.

_ICA_CON_BASE = ["codigo_municipio", "municipio", "concepto", "tarifa", "base_uvt"]


def test_la_base_minima_viaja_por_fila():
    repo = _FakeRepo()
    content = _xlsx(
        ica=[
            ["11001", "Bogotá D.C.", "servicios", 0.966, 4],
            ["11001", "Bogotá D.C.", "compras", 1.104, 27],
            ["68001", "Bucaramanga", "servicios", 0.7, 25],
        ],
        ica_headers=_ICA_CON_BASE,
    )

    ImportRetentionRatesUseCase(repo).execute(content)

    assert [r["minimum_base_uvt"] for r in repo.ica] == [4, 27, 25]


def test_sin_base_el_municipio_no_fija_tope():
    """Columna ausente o celda vacía = toda operación retiene, sin mínimo."""
    repo = _FakeRepo()
    content = _xlsx(ica=[["11001", "Bogotá D.C.", 0.966]])

    ImportRetentionRatesUseCase(repo).execute(content)

    assert repo.ica[0]["minimum_base_uvt"] is None


# ── Las líneas de ayuda de la plantilla no son datos ──────────────────────────


def test_las_filas_de_comentario_no_se_importan():
    """La plantilla lleva ejemplos precedidos de «#»; no pueden acabar en el catálogo.

    Antes solo se descartaba la fila completamente vacía, así que seguir la propia plantilla
    creaba un municipio llamado «# ejemplo» con tarifa 0,966.
    """
    repo = _FakeRepo()
    content = _xlsx(
        ica=[
            ["# ejemplo", "Bogotá D.C.", "servicios", 0.966, 4],
            ["# base_uvt = tope por debajo del cual no se retiene", None, None, None, None],
            ["11001", "Bogotá D.C.", "servicios", 0.966, 4],
        ],
        ica_headers=_ICA_CON_BASE,
    )

    result = ImportRetentionRatesUseCase(repo).execute(content)

    assert result["ica_loaded"] == 1
    assert repo.ica[0]["municipality_code"] == "11001"


def test_las_filas_de_comentario_tampoco_en_retefuente():
    repo = _FakeRepo()
    content = _xlsx(
        fuente=[
            ["# ejemplo de concepto", "declarante", 10, 523740, 2.5],
            ["Compras generales", "declarante", 10, 523740, 2.5],
        ]
    )

    result = ImportRetentionRatesUseCase(repo).execute(content)

    assert result["fuente_loaded"] == 1
    assert repo.fuente[0]["retention_concept"] == "Compras generales"


# ── La plantilla que se descarga ──────────────────────────────────────────────


def _plantilla() -> bytes:
    """Plantilla completa (las dos hojas), tal como se descargaba antes de separarlas.

    La hoja se pasa explícita: al invocar la función del endpoint directamente, sin pasar por
    FastAPI, un parámetro omitido llega como el objeto `Query(...)` y no como su valor por
    defecto.
    """
    return _plantilla_de(None)


def test_la_hoja_de_reteica_llega_vacia():
    """Una plantilla que hay que limpiar antes de usar no es una plantilla.

    Llevaba filas de ejemplo y líneas de ayuda dentro de la propia tabla: el contador tenía
    que borrar siete filas antes de escribir la primera suya. La ayuda vive ahora en la hoja
    de instrucciones y en las notas de cada cabecera.
    """
    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(_plantilla()))["ReteICA"]

    assert ws.max_row == 1  # solo la cabecera
    assert [c.value for c in ws[1]] == [
        "codigo_municipio",
        "municipio",
        "concepto",
        "tarifa",
        "base_uvt",
    ]


def test_cada_columna_explica_su_contenido_en_una_nota():
    """La ayuda se lee al pasar el cursor y no ocupa ninguna fila de datos."""
    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(_plantilla()))["ReteICA"]

    assert all(celda.comment is not None for celda in ws[1])
    assert "11001" in ws["A1"].comment.text  # el código DANE de Bogotá, a mano


def test_las_instrucciones_van_en_su_propia_hoja():
    """El importador solo lee ReteFuente y ReteICA, así que esa hoja se ignora sola."""
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(_plantilla()))

    assert "Instrucciones" in wb.sheetnames
    # Y no interfiere: la plantilla recién descargada se importa sin errores.
    repo = _FakeRepo()
    resultado = ImportRetentionRatesUseCase(repo).execute(_plantilla())
    assert resultado["ica_loaded"] == 0  # hoja vacía → no se toca la tabla
    assert resultado["fuente_loaded"] > 0  # ReteFuente sí viene precargada


# ── Importación acotada a una hoja ────────────────────────────────────────────
#
# La pantalla tiene un botón por tabla. Sin acotar, subir el archivo de una en el diálogo de
# la otra la cargaría en silencio; y en modo reemplazo habría vaciado antes la tabla que el
# usuario creía estar actualizando.


def test_sheet_ica_ignora_la_hoja_de_retefuente():
    repo = _FakeRepo()
    content = _xlsx(
        fuente=[["Compras generales", "declarante", 10, 523740, 2.5]],
        ica=[["11001", "Bogotá D.C.", "servicios", 0.966, 4]],
        ica_headers=_ICA_CON_BASE,
    )

    resultado = ImportRetentionRatesUseCase(repo).execute(content, sheet="ica")

    assert resultado == {"fuente_loaded": 0, "ica_loaded": 1}
    assert repo.fuente is None  # `None` = esa tabla no se toca en absoluto


def test_sheet_fuente_ignora_la_hoja_de_reteica():
    repo = _FakeRepo()
    content = _xlsx(
        fuente=[["Compras generales", "declarante", 10, 523740, 2.5]],
        ica=[["11001", "Bogotá D.C.", "servicios", 0.966, 4]],
        ica_headers=_ICA_CON_BASE,
    )

    resultado = ImportRetentionRatesUseCase(repo).execute(content, sheet="fuente")

    assert resultado == {"fuente_loaded": 1, "ica_loaded": 0}
    assert repo.ica is None


def test_el_archivo_equivocado_se_rechaza_con_un_mensaje_util():
    """Subir el archivo de ReteFuente en el diálogo de ReteICA no puede pasar inadvertido."""
    repo = _FakeRepo()
    content = _xlsx(fuente=[["Compras generales", "declarante", 10, 523740, 2.5]])

    with pytest.raises(ValidationException) as exc:
        ImportRetentionRatesUseCase(repo).execute(content, sheet="ica")

    assert "ReteICA" in str(exc.value)


def test_el_archivo_equivocado_no_borra_nada_en_modo_reemplazo():
    """La protección importa sobre todo aquí: reemplazar vacía antes de cargar."""
    repo = _FakeRepo()
    content = _xlsx(fuente=[["Compras generales", "declarante", 10, 523740, 2.5]])

    with pytest.raises(ValidationException):
        ImportRetentionRatesUseCase(repo).execute(content, replace=True, sheet="ica")

    assert repo.fuente is None and repo.ica is None  # no se llamó al repositorio


def test_una_hoja_no_reconocida_se_rechaza():
    repo = _FakeRepo()

    with pytest.raises(ValidationException) as exc:
        ImportRetentionRatesUseCase(repo).execute(_xlsx(ica=[["11001", "B", 1]]), sheet="iva")

    assert "fuente" in str(exc.value) and "ica" in str(exc.value)


def test_sin_sheet_se_procesan_las_dos_hojas():
    """Compatibilidad: el comportamiento anterior sigue disponible."""
    repo = _FakeRepo()
    content = _xlsx(
        fuente=[["Compras generales", "declarante", 10, 523740, 2.5]],
        ica=[["11001", "Bogotá D.C.", "servicios", 0.966, 4]],
        ica_headers=_ICA_CON_BASE,
    )

    assert ImportRetentionRatesUseCase(repo).execute(content) == {
        "fuente_loaded": 1,
        "ica_loaded": 1,
    }


# ── Plantilla por hoja ────────────────────────────────────────────────────────


def _plantilla_de(sheet):
    import asyncio

    from app.adapters.api.routers.catalog import download_retention_rates_template

    respuesta = download_retention_rates_template(sheet=sheet)

    async def _leer():
        return b"".join([trozo async for trozo in respuesta.body_iterator])

    return asyncio.run(_leer())


def test_cada_tabla_descarga_solo_su_hoja():
    """Entregar la hoja que no se va a llenar invita a subirla vacía."""
    from openpyxl import load_workbook

    fuente = load_workbook(BytesIO(_plantilla_de("fuente"))).sheetnames
    ica = load_workbook(BytesIO(_plantilla_de("ica"))).sheetnames

    assert "ReteFuente" in fuente and "ReteICA" not in fuente
    assert "ReteICA" in ica and "ReteFuente" not in ica
    # La ayuda viaja siempre, y el importador la ignora por no llamarse como las hojas de datos.
    assert "Instrucciones" in fuente and "Instrucciones" in ica


def test_la_plantilla_de_reteica_sigue_llegando_vacia():
    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(_plantilla_de("ica")))["ReteICA"]

    assert ws.max_row == 1
    assert all(celda.comment is not None for celda in ws[1])


def test_una_hoja_invalida_en_la_plantilla_da_400():
    from fastapi import HTTPException

    with pytest.raises(HTTPException) as exc:
        _plantilla_de("iva")

    assert exc.value.status_code == 400


# ── Códigos numéricos: la cola decimal rompe la identidad del municipio ───────


def test_un_codigo_escrito_como_numero_no_arrastra_el_decimal():
    """Excel, Numbers y LibreOffice guardan el código como número.

    openpyxl lo devuelve como float según cómo se escribiera el archivo, y un `str()` directo
    convertía el código DANE 11001 en '11001.0'. No es cosmético: con el decimal pegado no
    cruza con '11001' en la restricción única, ni en el filtro por municipio con el que RF-08
    recupera casos contabilizados, ni contra ningún catálogo.
    """
    caso = ImportRetentionRatesUseCase(_FakeRepo())

    assert caso._text([11001.0], 0) == "11001"
    assert caso._text([11001], 0) == "11001"


def test_un_decimal_real_conserva_sus_decimales():
    """Solo se normaliza el entero: perder la parte decimal en silencio sería peor."""
    assert ImportRetentionRatesUseCase(_FakeRepo())._text([0.966], 0) == "0.966"


def test_el_codigo_llega_limpio_a_la_fila_importada():
    repo = _FakeRepo()
    content = _xlsx(
        ica=[[11001.0, "Bogotá D.C.", "Servicios", 0.966, 4]],
        ica_headers=_ICA_CON_BASE,
    )

    ImportRetentionRatesUseCase(repo).execute(content, sheet="ica")

    assert repo.ica[0]["municipality_code"] == "11001"


# ── El endpoint debe devolver TODAS las columnas del DTO ─────────────────────


class _FilaIca:
    """Fila del ORM tal como la entrega el repositorio."""

    municipality_code = "11001"
    municipality_name = "Bogotá D.C."
    retention_concept = "servicios"
    percentage = 0.966
    minimum_base_uvt = 4


class _RepoIca:
    def get_ica_rates(self):
        return [_FilaIca()]


def test_el_endpoint_no_pierde_columnas_por_el_camino():
    """El endpoint construía el DTO campo a campo y olvidaba los nuevos, sin fallar.

    Los campos no pasados caían a su valor por defecto, así que la interfaz mostraba
    «Todos los conceptos» y «Sin tope» mientras la base tenía 'servicios' y 4 UVT. Un dato
    correcto mostrado mal es peor que un error: nadie lo investiga.

    Esta prueba compara contra los campos declarados en el DTO, así que cubre también las
    columnas que se añadan en el futuro.
    """
    from app.adapters.api.routers.catalog import get_retention_ica_rates
    from app.application.dto.catalog import RetentionIcaRateResponse

    resultado = get_retention_ica_rates(repo=_RepoIca())[0].model_dump()

    assert set(resultado) == set(RetentionIcaRateResponse.model_fields)
    # Y con los valores de la fila, no con los de por defecto del DTO.
    assert resultado["retention_concept"] == "servicios"
    assert resultado["minimum_base_uvt"] == 4
