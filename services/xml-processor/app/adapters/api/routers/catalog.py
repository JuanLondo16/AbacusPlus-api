from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from app.application.dto.catalog import (
    CostCenterResponse,
    ImportRetentionRatesResponse,
    PucAccountResponse,
    RetentionFuenteRateResponse,
    RetentionIcaRateResponse,
    TaxCatalogResponse,
)
from app.application.use_cases.import_retention_rates import ImportRetentionRatesUseCase
from app.dependencies import (
    get_cost_center_repo,
    get_import_retention_rates_use_case,
    get_integration_retention_repo,
    get_integration_tax_repo,
    get_puc_repo,
    get_retention_repo,
)
from app.domain.value_objects.retention_scope import (
    es_impuesto_de_linea,
    es_retencion_practicable,
)
from app.infrastructure.config.auth_dependency import require_write
from app.infrastructure.persistence.repositories.cost_center_repository import CostCenterRepository
from app.infrastructure.persistence.repositories.integration_retention_repository import (
    IntegrationRetentionRepository,
)
from app.infrastructure.persistence.repositories.integration_tax_repository import (
    IntegrationTaxRepository,
)
from app.infrastructure.persistence.repositories.puc_repository import PucRepository
from app.infrastructure.persistence.repositories.retention_repository import RetentionRepository

router = APIRouter()


@router.get(
    "/catalog/cost-centers",
    response_model=list[CostCenterResponse],
    summary="Listar centros de costo activos",
    description=(
        "Retorna todos los centros de costo activos configurados en el sistema. "
        "El llm-service los usa para que el LLM asigne centros de costo reales "
        "(en lugar de dejar `null`) al generar asientos contables de causación."
    ),
    response_description="Lista de centros de costo con código y nombre.",
)
def get_cost_centers(
    repo: CostCenterRepository = Depends(get_cost_center_repo),
):
    return [CostCenterResponse.model_validate(cc) for cc in repo.get_active()]


@router.get(
    "/catalog/taxes",
    response_model=list[TaxCatalogResponse],
    summary="Listar impuestos/retenciones del catálogo",
    description=(
        "RF-02: retorna los impuestos y retenciones activos del catálogo (id, nombre, tipo y "
        "porcentaje).\n\n"
        "Desde la migración del 2026-08-31 el catálogo vive en DOS tablas físicas: "
        "`integration_taxes` (impuestos reales del documento: IVA, Impoconsumo, AdValorem) e "
        "`integration_retentions` (retenciones: ReteICA, ReteIVA, Retefuente, "
        "Autorretención). `ambito` decide de cuál — o de las dos — se lee.\n\n"
        "Alimenta el selector de la sección de retenciones del detalle del documento, de modo "
        "que el usuario solo pueda agregar retenciones que existen en el catálogo."
    ),
    response_description="Lista de impuestos/retenciones con id, nombre, tipo y porcentaje.",
)
def get_taxes(
    ambito: str = Query(
        "retenciones",
        pattern="^(retenciones|linea|todos)$",
        description=(
            "Qué parte del catálogo se necesita. SIIGO reparte los impuestos en dos sitios "
            "del comprobante y cada selector debe ofrecer solo los suyos:\n\n"
            "- `retenciones` (por defecto): las que SIIGO practica a nivel de documento "
            "(ReteICA, ReteIVA), leídas de `integration_retentions`. Cada fila `reteica` trae "
            "ya su municipio, concepto y base mínima. Alimenta el selector de retenciones.\n"
            "- `linea`: los impuestos que se asignan a un ítem y suman al valor de la "
            "operación (IVA, Impoconsumo, AdValorem), leídos de `integration_taxes`. Alimenta "
            "el selector de la línea.\n"
            "- `todos`: el catálogo activo completo de las DOS tablas. Solo para resolver el "
            "nombre de lo ya registrado (`document_taxes.tax_id` puede apuntar a cualquiera "
            "de las dos), incluido lo que hoy no se ofrecería."
        ),
    ),
    repo: IntegrationTaxRepository = Depends(get_integration_tax_repo),
    retention_repo: IntegrationRetentionRepository = Depends(get_integration_retention_repo),
):
    # Un tipo no puede acabar en el sitio equivocado: un impuesto puesto donde va una
    # retención se resta en lugar de sumarse, que es lo que descuadró cuatro documentos con
    # el impuesto al consumo. Por eso cada ámbito filtra con la regla del dominio y no con
    # una lista escrita aquí.
    if ambito == "linea":
        return [
            TaxCatalogResponse.model_validate(t)
            for t in repo.get_active()
            if es_impuesto_de_linea(getattr(t, "type", None))
        ]

    if ambito == "todos":
        # Sin filtro adicional de tipo: solo sirve para poner nombre a lo ya registrado,
        # nunca para ofrecerlo. Combina las dos tablas porque un `tax_id` guardado puede
        # resolver en cualquiera de las dos desde la separación del 2026-08-31.
        return [
            TaxCatalogResponse.model_validate(t)
            for t in [*repo.get_active(), *retention_repo.get_active()]
        ]

    # ambito == "retenciones" (por defecto): solo lo que SIIGO puede practicar en una factura
    # de compra (ReteICA, ReteIVA — ver `es_retencion_practicable`), leído de
    # `integration_retentions`.
    #
    # Antes se devolvía el catálogo entero de `integration_taxes` (que entonces mezclaba
    # impuestos y retenciones), así que el selector ofrecía también Retefuente, Autorretención
    # e Impoconsumo. El contador las registraba, Abacus las descontaba del total a pagar y el
    # envío las descartaba porque la API las rechaza: el documento acababa contabilizado por
    # el importe íntegro y la pantalla mostraba otro. No ofrecerlas es lo que evita esa
    # diferencia, en lugar de explicarla después.
    return [
        TaxCatalogResponse.model_validate(t)
        for t in retention_repo.get_active()
        if es_retencion_practicable(getattr(t, "type", None))
    ]


@router.get(
    "/catalog/puc-accounts",
    response_model=list[PucAccountResponse],
    summary="Listar cuentas PUC activas",
    description=(
        "Retorna todas las cuentas del Plan Único de Cuentas (PUC) marcadas como activas. "
        "El llm-service las inyecta en el contexto del LLM como catálogo de cuentas válidas, "
        "reduciendo la probabilidad de que el modelo invente códigos de cuenta inexistentes."
    ),
    response_description="Lista de cuentas PUC activas con código, nombre y nivel.",
)
def get_puc_accounts(
    repo: PucRepository = Depends(get_puc_repo),
):
    return [PucAccountResponse.model_validate(a) for a in repo.get_active()]


@router.get(
    "/catalog/retention-fuente-rates",
    response_model=list[RetentionFuenteRateResponse],
    summary="Listar tasas de retención en la fuente",
    description=(
        "Retorna las tasas de retención en la fuente configuradas por concepto y contribuyente. "
        "El llm-service las usa como referencia para determinar la subcuenta 2365xx correcta "
        "según el tipo de proveedor y bases mínimas (pesos y UVT). "
    ),
    response_description="Lista de tasas de reteFuente por concepto y contribuyente.",
)
def get_retention_fuente_rates(
    repo: RetentionRepository = Depends(get_retention_repo),
):
    # Igual que en ReteICA: el mapeo lo define el DTO, no una lista de campos escrita a mano
    # que hay que mantener en paralelo. Aquí estaba completa, pero era el mismo defecto
    # esperando a la siguiente columna.
    return [RetentionFuenteRateResponse.model_validate(r) for r in repo.get_fuente_rates()]


@router.get(
    "/catalog/retention-ica-rates",
    response_model=list[RetentionIcaRateResponse],
    summary="Listar tasas de retención ICA por municipio",
    description=(
        "Retorna las tasas de reteICA configuradas por municipio (código DANE). "
        "Referencia para que el LLM calcule o valide el valor correcto de reteICA "
        "al generar asientos de causación de facturas locales."
    ),
    response_description="Lista de tasas de reteICA por código y nombre de municipio.",
)
def get_retention_ica_rates(
    repo: RetentionRepository = Depends(get_retention_repo),
):
    # `model_validate` sobre la fila del ORM y no un constructor campo a campo.
    #
    # El constructor manual dejaba fuera cualquier columna nueva **sin fallar**: los campos
    # que no se pasaban caían a su valor por defecto, así que `retention_concept` llegaba a la
    # interfaz como 'todos' y `minimum_base_uvt` como None aunque la base tuviera el dato
    # correcto. Un dato correcto mostrado mal es peor que un error: nadie lo investiga.
    #
    # Con `from_attributes` el mapeo lo define el DTO. Añadir una columna al modelo y al DTO
    # basta para que viaje; no hay un tercer sitio que recordar.
    return [RetentionIcaRateResponse.model_validate(r) for r in repo.get_ica_rates()]


@router.post(
    "/catalog/retention-rates/imports",
    dependencies=[Depends(require_write)],
    response_model=ImportRetentionRatesResponse,
    summary="Importar tarifas de retención (ReteFuente / ReteICA) desde Excel",
    description=(
        "Recibe un archivo `.xlsx` con las tarifas de retención y las carga en la base del "
        "tenant, para que persistan y el contador pueda actualizarlas cuando cambien (UVT y "
        "tarifas anuales, ReteICA por municipio).\n\n"
        "El archivo puede traer una o ambas hojas:\n"
        "- **`ReteFuente`** (columnas: `concepto`, `tipo_contribuyente`, `base_uvt`, "
        "`base_pesos`, `tarifa`).\n"
        "- **`ReteICA`** (columnas: `codigo_municipio`, `municipio`, `concepto`, `tarifa`, "
        "`base_uvt`).\n\n"
        "  Un municipio puede traer **varias filas, una por concepto** (compra, servicios, "
        "honorarios, comisiones…), porque la tarifa de ReteICA la fija la actividad. La "
        "columna `concepto` es opcional: si no viene, la tarifa se guarda como `todos` y "
        "aplica a cualquier operación de ese municipio.\n\n"
        "  `base_uvt` es el tope por debajo del cual NO se retiene, **en UVT**. Lo fija cada "
        "municipio y no hay uniformidad nacional (Bogotá 4/27, Cali 3/15, Bucaramanga 25/50), "
        "por eso viaja en la tabla y no en el código. Se guarda en UVT y no en pesos porque la "
        "DIAN actualiza la UVT cada año: la conversión se hace con la del año del documento. "
        "Vacía = el municipio no fija tope para ese concepto.\n\n"
        "**Qué se hace con lo ya cargado** lo decide `replace`, con la misma semántica que la "
        "importación del plan de cuentas:\n"
        "- `false` (por defecto): **upsert** — cada fila corrige o agrega, y lo que no venga "
        "en el archivo se conserva. Es el modo seguro para ajustar unos pocos conceptos.\n"
        "- `true`: **reemplazo** — se vacía la tabla antes de cargar, de modo que el archivo "
        "pasa a ser la verdad completa. Es lo que se necesita en la re-importación anual, "
        "cuando además hay que dar de baja conceptos que dejaron de existir.\n\n"
        "Una hoja ausente **nunca se toca**, ni siquiera con `replace=true`: subir solo "
        "ReteICA no puede borrar la tabla nacional.\n\n"
        "`tipo_contribuyente` válido: declarante · no_declarante · todos · personas_juridicas · "
        "personas_naturales."
    ),
    response_description="Cantidad de tarifas cargadas por tipo.",
    responses={
        400: {"description": "Archivo inválido, hojas/columnas faltantes o datos incorrectos."}
    },
)
async def import_retention_rates(
    file: UploadFile = File(
        ..., description="Archivo Excel .xlsx con hojas ReteFuente y/o ReteICA."
    ),
    replace: bool = Form(
        False,
        description=(
            "`false`: actualiza las tarifas existentes y agrega las nuevas, sin eliminar. "
            "`true`: elimina las tarifas actuales de las hojas incluidas antes de importar."
        ),
        examples=[False],
    ),
    sheet: Optional[str] = Form(
        None,
        description=(
            "Acota la importación a una sola hoja: `fuente` o `ica`. La interfaz lo envía "
            "porque cada tabla tiene su propio botón de importar, y sin acotar, subir el "
            "archivo de ReteICA en el diálogo de ReteFuente cargaría la hoja equivocada en "
            "silencio. Si el archivo no trae la hoja pedida, la importación se rechaza. "
            "Omitir el parámetro procesa las dos hojas."
        ),
        examples=["ica"],
    ),
    use_case: ImportRetentionRatesUseCase = Depends(get_import_retention_rates_use_case),
):
    content = await file.read()
    result = use_case.execute(file_content=content, replace=replace, sheet=sheet)
    return ImportRetentionRatesResponse(**result)


# ── Plantilla de importación ───────────────────────────────────────────────────
#
# La plantilla se genera por hoja porque cada tabla de la interfaz tiene su propia
# importación. Entregar siempre las dos hojas invitaba a subir vacía la que no se iba a
# llenar, y una hoja vacía no es inofensiva: en modo reemplazo habría borrado esa tabla.

_CABECERA_FONDO = "1565C0"

#: Qué explica cada columna. Va como nota de Excel (el triangulito rojo de la celda), que se
#: lee al pasar el cursor y no ocupa ninguna fila de datos: la plantilla llega limpia.
_NOTAS_FUENTE = {
    "concepto": (
        "Concepto tributario de la operación. Junto con el tipo de contribuyente identifica "
        "la fila de forma única."
    ),
    "tipo_contribuyente": (
        "declarante · no_declarante · todos · personas_juridicas · personas_naturales"
    ),
    "base_uvt": "Tope en UVT por debajo del cual no se retiene. Vacío = sin tope.",
    "base_pesos": "Equivalente en pesos de la base, para referencia del contador.",
    "tarifa": "Porcentaje. Escriba 2,5 para un 2,5 %.",
}

_NOTAS_ICA = {
    "codigo_municipio": (
        "Código DANE del municipio. Bogotá D.C. = 11001, Medellín = 05001, Cali = 76001."
    ),
    "municipio": "Nombre del municipio, para poder leer la tabla.",
    "concepto": (
        "Actividad que fija la tarifa: servicios, compras, honorarios, comisiones…\n"
        "Un municipio puede llevar varias filas, una por concepto.\n"
        "Use 'todos' si aplica una sola tarifa a cualquier operación."
    ),
    "tarifa": "Porcentaje. Escriba 0,966 para un 0,966 %.",
    "base_uvt": (
        "Tope en UVT por debajo del cual NO se retiene. Lo fija cada municipio:\n"
        "Bogotá 4/27 · Cali 3/15 · Medellín 15 · Bucaramanga 25/50 (servicios/compras).\n"
        "Déjelo vacío si su municipio no fija tope para ese concepto."
    ),
}


def _formatear_hoja(ws, anchos: dict, notas: dict) -> None:
    """Cabecera destacada y fija, columnas legibles y una nota por columna."""
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill

    for celda in ws[1]:
        celda.fill = PatternFill("solid", fgColor=_CABECERA_FONDO)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center", vertical="center")
        if celda.value in notas:
            celda.comment = Comment(notas[celda.value], "Abacus", height=150, width=340)
    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho
    ws.freeze_panes = "A2"  # la cabecera sigue visible al desplazarse


def _hoja_instrucciones(wb, sheet) -> None:
    """Hoja de ayuda, adaptada a la plantilla que se está descargando.

    Va en su propia hoja porque el importador solo lee «ReteFuente» y «ReteICA» por nombre:
    cualquier otra se ignora sin riesgo, y así la ayuda no invade los datos.
    """
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet("Instrucciones")
    ws.column_dimensions["A"].width = 105
    titulo = Font(bold=True, size=13, color=_CABECERA_FONDO)
    subtitulo = Font(bold=True)
    fila = [0]

    def escribir(texto: str = "", fuente=None) -> None:
        fila[0] += 1
        celda = ws.cell(row=fila[0], column=1, value=texto)
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        if fuente:
            celda.font = fuente

    escribir("Tarifas de retención · cómo llenar esta plantilla", titulo)
    escribir()
    escribir("Al importar puede elegir entre dos modos:")
    escribir(
        "  · Actualizar (por defecto): cada fila corrige o agrega; lo que no venga se conserva."
    )
    escribir("  · Reemplazar: el archivo pasa a ser la verdad completa y se borra lo anterior.")
    escribir("Esta hoja de instrucciones se ignora al importar.")
    escribir()

    if sheet in (None, "fuente"):
        escribir("Hoja ReteFuente — tarifas nacionales", subtitulo)
        escribir("Viene precargada con la tabla vigente. Revísela con su contador y ajuste lo")
        escribir("que corresponda. Cada fila se identifica por concepto + tipo_contribuyente.")
        escribir()

    if sheet in (None, "ica"):
        escribir("Hoja ReteICA — tarifas municipales", subtitulo)
        escribir("Llega vacía porque el ICA es un impuesto territorial: no hay tabla nacional")
        escribir("que precargar. Agregue una fila por cada combinación de municipio y concepto")
        escribir("donde su empresa retiene. Los municipios que aparezcan aquí son los únicos en")
        escribir("los que el sistema sugerirá ReteICA.")
        escribir()
        escribir("Un municipio puede llevar varias filas, una por concepto, porque la tarifa la")
        escribir("fija la actividad. Por ejemplo, en Bogotá:")
        escribir("     11001 · Bogotá D.C. · servicios · 0,966 · 4")
        escribir("     11001 · Bogotá D.C. · compras   · 1,104 · 27")
        escribir()
        escribir("Escriba los conceptos con la nomenclatura de su municipio. Cuanto más se")
        escribir("parezcan a las descripciones que llegan en las facturas, mejor las emparejará")
        escribir("el sistema. Use «todos» si aplica una sola tarifa a cualquier operación.")
        escribir()
        escribir("Base mínima (base_uvt)", subtitulo)
        escribir("Tope en UVT por debajo del cual NO se practica la retención. Lo fija cada")
        escribir("municipio y no hay uniformidad nacional:")
        escribir("     Bogotá 4 / 27      Cali 3 / 15      Medellín 15      Bucaramanga 25 / 50")
        escribir("     (servicios / compras)")
        escribir("Se indica en UVT, no en pesos: la DIAN actualiza el valor de la UVT cada año y")
        escribir("el sistema hace la conversión con la del año de cada factura.")
        escribir("Déjela vacía si su municipio no fija tope para ese concepto.")


def _hoja_retefuente(wb) -> None:
    """Cabecera con formato y la tabla nacional vigente precargada, para revisarla."""
    from app.domain.services.retention_fuente_seed import STANDARD_RETEFUENTE_2026

    ws = wb.create_sheet("ReteFuente")
    ws.append(["concepto", "tipo_contribuyente", "base_uvt", "base_pesos", "tarifa"])
    for fila in STANDARD_RETEFUENTE_2026:
        ws.append(
            [
                fila["retention_concept"],
                fila["taxpayer_type"],
                fila["minimum_base_uvt"],
                fila["minimum_base_pesos"],
                fila["rate_percentage"],
            ]
        )
    _formatear_hoja(ws, {"A": 52, "B": 20, "C": 12, "D": 14, "E": 10}, _NOTAS_FUENTE)


def _hoja_reteica(wb) -> None:
    """Solo la cabecera: las tarifas municipales las carga el contador.

    Se deja sin filas de ejemplo a propósito. Una plantilla que hay que limpiar antes de
    usarla no es una plantilla, y una tarifa de ejemplo importada por descuido es una
    retención mal calculada.
    """
    ws = wb.create_sheet("ReteICA")
    ws.append(["codigo_municipio", "municipio", "concepto", "tarifa", "base_uvt"])
    _formatear_hoja(ws, {"A": 18, "B": 26, "C": 24, "D": 10, "E": 12}, _NOTAS_ICA)


@router.get(
    "/catalog/retention-rates/template",
    summary="Descargar plantilla Excel de tarifas de retención",
    description=(
        "Genera un `.xlsx` listo para llenar y volver a importar.\n\n"
        "Con `sheet=fuente` trae solo la hoja **`ReteFuente`**, precargada con la tabla "
        "nacional vigente para que el contador ajuste tarifas y bases del año.\n\n"
        "Con `sheet=ica` trae solo la hoja **`ReteICA`**, con los encabezados y nada más: "
        "las tarifas municipales no tienen tabla nacional que precargar.\n\n"
        "Sin `sheet` devuelve ambas. En todos los casos se incluye una hoja "
        "**`Instrucciones`**, que el importador ignora."
    ),
    response_description="Archivo .xlsx de plantilla.",
    responses={400: {"description": "El valor de `sheet` no es `fuente` ni `ica`."}},
)
def download_retention_rates_template(
    sheet: Optional[str] = Query(
        None,
        description=(
            "Hoja a incluir: `fuente` o `ica`. Cada tabla de la interfaz descarga la suya, "
            "para no entregar una hoja que no se va a llenar. Omitir devuelve ambas."
        ),
        examples=["ica"],
    ),
):
    """Plantilla .xlsx lista para rellenar.

    La hoja de datos llega **limpia**: cabecera y nada más (salvo ReteFuente, cuya tabla
    nacional sí es un punto de partida real). La ayuda vive donde Excel espera que viva —una
    hoja de instrucciones y una nota por columna—, no mezclada con las filas.
    """
    from openpyxl import Workbook

    if sheet is not None and sheet not in ("fuente", "ica"):
        raise HTTPException(
            status_code=400,
            detail="El parámetro 'sheet' solo admite 'fuente' o 'ica'.",
        )

    wb = Workbook()
    wb.remove(wb.active)

    _hoja_instrucciones(wb, sheet)
    if sheet in (None, "fuente"):
        _hoja_retefuente(wb)
    if sheet in (None, "ica"):
        _hoja_reteica(wb)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    sufijo = {"fuente": "-retefuente", "ica": "-reteica"}.get(sheet or "", "")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="plantilla-tarifas{sufijo}.xlsx"'},
    )
