"""Importación de tarifas de retención (ReteFuente y ReteICA) desde un archivo Excel.

Motivación (RF-08): las tarifas de retención son datos de referencia que cambian con el tiempo
—UVT y tarifas anuales, tarifas de ReteICA por municipio— y deben persistir por empresa cuando
la aplicación se entrega a otro cliente. En vez de vivir en el código, se cargan en la base de
cada tenant (`retention_fuente_rates`, `retention_ica_rates`) y el contador las actualiza
subiendo una plantilla `.xlsx`, con el mismo patrón de importación que el PUC y los impuestos.

El archivo puede traer una hoja `ReteFuente`, una hoja `ReteICA`, o ambas. Qué se hace con lo
ya cargado lo decide `replace`, con la misma semántica que la importación del plan de cuentas:

- `replace=False` (por defecto): upsert. Cada fila corrige o agrega, y lo que no venga en el
  archivo se conserva.
- `replace=True`: se vacía la tabla antes de cargar; el archivo pasa a ser la verdad completa.

Una hoja ausente nunca se toca, ni siquiera con `replace=True`: subir solo ReteICA no puede
borrar la tabla nacional.
"""

import unicodedata
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook

from app.domain.exceptions.base import ValidationException
from app.domain.services.ica_rate_units import (
    UnidadesMezcladasError,
    verificar_unidad_coherente,
)
from app.infrastructure.persistence.models.retention_ica import CONCEPTO_GENERAL
from app.infrastructure.persistence.repositories.retention_repository import RetentionRepository

# Tipos de contribuyente admitidos en ReteFuente (mismos que usa suggest_retentions).
_VALID_TAXPAYER_TYPES = {
    "declarante",
    "no_declarante",
    "todos",
    "personas_juridicas",
    "personas_naturales",
}

# Mapa de encabezado normalizado -> campo del modelo. Se aceptan sinónimos en español.
_FUENTE_HEADERS = {
    "concepto": "retention_concept",
    "retention_concept": "retention_concept",
    "tipo_contribuyente": "taxpayer_type",
    "tipo": "taxpayer_type",
    "taxpayer_type": "taxpayer_type",
    "base_uvt": "minimum_base_uvt",
    "minimum_base_uvt": "minimum_base_uvt",
    "base_pesos": "minimum_base_pesos",
    "minimum_base_pesos": "minimum_base_pesos",
    "tarifa": "rate_percentage",
    "tarifa_porcentaje": "rate_percentage",
    "porcentaje": "rate_percentage",
    "rate_percentage": "rate_percentage",
}
_ICA_HEADERS = {
    "codigo_municipio": "municipality_code",
    "codigo": "municipality_code",
    "municipality_code": "municipality_code",
    "municipio": "municipality_name",
    "nombre_municipio": "municipality_name",
    "municipality_name": "municipality_name",
    "tarifa": "percentage",
    "porcentaje": "percentage",
    "percentage": "percentage",
    # El concepto determina la tarifa dentro de un mismo municipio (compra, servicios,
    # honorarios…). Es opcional: una plantilla antigua sin esta columna sigue siendo válida y
    # sus filas quedan como 'todos'.
    "concepto": "retention_concept",
    "concepto_retencion": "retention_concept",
    "actividad": "retention_concept",
    "retention_concept": "retention_concept",
    # Base mínima del municipio, en UVT. Opcional: sin ella no se aplica tope y toda
    # operación retiene, que es como se comportaba antes de existir la columna.
    "base_uvt": "minimum_base_uvt",
    "base_minima_uvt": "minimum_base_uvt",
    "minimum_base_uvt": "minimum_base_uvt",
}


class ImportRetentionRatesUseCase:
    FUENTE_SHEET = "ReteFuente"
    ICA_SHEET = "ReteICA"

    def __init__(self, repository: RetentionRepository):
        self._repo = repository

    #: Hojas que puede pedir el llamador. `None` = ambas (compatibilidad y carga conjunta).
    SHEET_KEYS = {"fuente": "FUENTE_SHEET", "ica": "ICA_SHEET"}

    def execute(
        self,
        file_content: bytes,
        replace: bool = False,
        sheet: Optional[str] = None,
    ) -> dict[str, int]:
        """Importa las tarifas del archivo. `sheet` acota la importación a una sola hoja.

        Acotar no es cosmético: la interfaz tiene un botón por tabla, y sin este parámetro
        subir el archivo de ReteICA en el diálogo de ReteFuente importaría la hoja equivocada
        **en silencio**, con el agravante de que en modo reemplazo habría vaciado antes la
        tabla que el usuario creía estar cargando. Con `sheet`, cualquier archivo que no
        traiga la hoja pedida se rechaza con un mensaje que dice exactamente qué falta.
        """
        if sheet is not None and sheet not in self.SHEET_KEYS:
            raise ValidationException(
                f"Hoja '{sheet}' no reconocida. Valores admitidos: "
                f"{', '.join(sorted(self.SHEET_KEYS))}."
            )

        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
        except Exception as exc:  # openpyxl lanza tipos variados ante archivos corruptos
            raise ValidationException("El archivo debe ser un Excel .xlsx válido") from exc

        sheets = {name.strip().lower(): name for name in workbook.sheetnames}
        fuente_rows = None
        ica_rows = None

        # Una hoja presente pero sin filas con datos (p. ej. la plantilla con ReteICA aún sin
        # llenar) se trata como «no provista» (None): así no se vacía por error esa tabla y la
        # importación del resto sigue. `or None` convierte la lista vacía en None.
        quiere_fuente = sheet in (None, "fuente")
        quiere_ica = sheet in (None, "ica")

        if quiere_fuente and self.FUENTE_SHEET.lower() in sheets:
            fuente_rows = self._parse_fuente(workbook[sheets[self.FUENTE_SHEET.lower()]]) or None
        if quiere_ica and self.ICA_SHEET.lower() in sheets:
            ica_rows = self._parse_ica(workbook[sheets[self.ICA_SHEET.lower()]]) or None

        # La tabla de ReteICA no puede mezclar unidades. El ICA se publica por mil —«9,66 por
        # mil» es 0,966 %— y las dos formas circulan en los documentos que maneja un contador,
        # así que un Excel puede traer las dos sin que nadie lo note. Ocurrió: la tabla del
        # cliente tenía 9.66 (por mil) y 1.104 (porcentaje) a la vez.
        #
        # Se comprueba ANTES de escribir. Detectarlo al sugerir, como se hacía, es tarde y
        # depende de que alguien pida una sugerencia; mientras tanto el dato incoherente vive
        # en la base y cualquier otra vía lo usará.
        if ica_rows:
            try:
                verificar_unidad_coherente(ica_rows)
            except UnidadesMezcladasError as exc:
                raise ValidationException(str(exc)) from exc

        if fuente_rows is None and ica_rows is None:
            esperada = (
                self.FUENTE_SHEET
                if sheet == "fuente"
                else self.ICA_SHEET
                if sheet == "ica"
                else f"'{self.FUENTE_SHEET}' o '{self.ICA_SHEET}'"
            )
            nombre = esperada if sheet is None else f"'{esperada}'"
            raise ValidationException(
                f"El archivo no tiene datos para importar. Debe traer una hoja {nombre} "
                f"con al menos una fila."
            )

        fuente_count, ica_count = self._repo.import_rates(fuente_rows, ica_rows, replace=replace)
        return {"fuente_loaded": fuente_count, "ica_loaded": ica_count}

    # ── Parseo por hoja ────────────────────────────────────────────────────────
    def _parse_fuente(self, sheet) -> list[dict[str, Any]]:
        header_map = self._header_map(
            sheet, _FUENTE_HEADERS, {"retention_concept", "taxpayer_type", "rate_percentage"}
        )
        rows: list[dict[str, Any]] = []
        # (concepto, tipo_contribuyente) es la clave única de la tabla. Dos filas iguales en el
        # archivo son un error del usuario y hay que decírselo: en reemplazo la inserción
        # fallaría contra la restricción única con un mensaje incomprensible, y en upsert la
        # segunda pisaría a la primera en silencio, cargando una tarifa que nadie eligió.
        seen: set[tuple[str, str]] = set()
        for line, raw in self._data_rows(sheet):
            concepto = self._text(raw, header_map.get("retention_concept"))
            if not concepto or self._es_comentario(concepto):
                continue  # fila vacía o anotación de la plantilla → se ignora
            taxpayer = (
                (self._text(raw, header_map.get("taxpayer_type")) or "").lower().replace(" ", "_")
            )
            if taxpayer not in _VALID_TAXPAYER_TYPES:
                raise ValidationException(
                    f"ReteFuente fila {line}: tipo_contribuyente inválido '{taxpayer}'. "
                    f"Válidos: {', '.join(sorted(_VALID_TAXPAYER_TYPES))}"
                )
            clave = (concepto, taxpayer)
            if clave in seen:
                raise ValidationException(
                    f"ReteFuente fila {line}: el concepto '{concepto}' ya aparece para el tipo "
                    f"de contribuyente '{taxpayer}'. Deje una sola fila por combinación."
                )
            seen.add(clave)
            rows.append(
                {
                    "retention_concept": concepto,
                    "taxpayer_type": taxpayer,
                    "minimum_base_uvt": self._num(
                        raw, header_map.get("minimum_base_uvt"), line, "base_uvt", required=False
                    ),
                    "minimum_base_pesos": self._num(
                        raw,
                        header_map.get("minimum_base_pesos"),
                        line,
                        "base_pesos",
                        required=False,
                    ),
                    "rate_percentage": self._num(
                        raw, header_map.get("rate_percentage"), line, "tarifa", required=True
                    ),
                }
            )
        return rows

    def _parse_ica(self, sheet) -> list[dict[str, Any]]:
        """Filas de ReteICA. La clave es (municipio, concepto), no el municipio solo.

        Un municipio tiene varias tarifas —una por concepto o actividad: compra, servicios,
        honorarios, comisiones—, así que repetir el código NO es un error: es lo normal. Lo
        que sí es un error es repetir el mismo par (municipio, concepto), porque entonces hay
        dos tarifas para el mismo caso y el archivo no dice cuál vale.

        La columna `concepto` es opcional. Sin ella, la fila queda como `todos`, que es lo
        que significaba una tarifa por municipio antes de que el concepto existiera: así una
        plantilla antigua se sigue importando sin cambios y sin reinterpretar su contenido.
        """
        header_map = self._header_map(sheet, _ICA_HEADERS, {"municipality_code", "percentage"})
        rows: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()
        for line, raw in self._data_rows(sheet):
            code = self._text(raw, header_map.get("municipality_code"))
            if not code or self._es_comentario(code):
                continue
            concepto = (
                (self._text(raw, header_map.get("retention_concept")) or CONCEPTO_GENERAL)
                .strip()
                .lower()
            )
            clave = (code, concepto)
            if clave in seen:
                raise ValidationException(
                    f"ReteICA fila {line}: ya hay una tarifa para el municipio '{code}' y el "
                    f"concepto '{concepto}'. Un municipio puede tener varias tarifas, pero "
                    f"solo una por concepto."
                )
            seen.add(clave)
            rows.append(
                {
                    "municipality_code": code,
                    "municipality_name": self._text(raw, header_map.get("municipality_name")),
                    "retention_concept": concepto,
                    "percentage": self._num(
                        raw, header_map.get("percentage"), line, "tarifa", required=True
                    ),
                    # Opcional: sin base, el municipio no fija tope para ese concepto.
                    "minimum_base_uvt": self._num(
                        raw, header_map.get("minimum_base_uvt"), line, "base_uvt", required=False
                    ),
                }
            )
        return rows

    # ── Helpers ────────────────────────────────────────────────────────────────
    def _header_map(self, sheet, alias: dict[str, str], required: set[str]) -> dict[str, int]:
        rows = sheet.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise ValidationException(f"La hoja '{sheet.title}' está vacía") from exc
        self._rows_iter = rows  # se reutiliza en _data_rows
        mapping: dict[str, int] = {}
        for idx, value in enumerate(header):
            key = self._normalize(value)
            field = alias.get(key)
            if field and field not in mapping:
                mapping[field] = idx
        missing = required - set(mapping)
        if missing:
            human = ", ".join(sorted(missing))
            raise ValidationException(
                f"La hoja '{sheet.title}' no tiene las columnas requeridas: {human}"
            )
        return mapping

    @staticmethod
    def _es_comentario(valor) -> bool:
        """True si la fila es una anotación de la plantilla, no un dato.

        Las líneas de ayuda de la plantilla empiezan por «#». Sin esta comprobación se
        importaban como si fueran datos —un municipio llamado «# ejemplo:» con tarifa 0,966—,
        porque la única fila que se descartaba era la que venía completamente vacía. El
        contador acababa con basura en su catálogo por seguir la propia plantilla.

        También sirve para las notas que el propio contador escriba en el archivo.
        """
        return str(valor or "").strip().startswith("#")

    def _data_rows(self, sheet):
        for i, row in enumerate(self._rows_iter, start=2):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            yield i, row

    @staticmethod
    def _normalize(value: Any) -> str:
        if value is None:
            return ""
        text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
        return text.strip().lower().replace(" ", "_")

    @staticmethod
    def _text(row, index: Optional[int]) -> Optional[str]:
        """Valor de una celda como texto, sin la cola decimal de los números enteros.

        Excel, Numbers y LibreOffice guardan una celda numérica como número, y openpyxl la
        devuelve como `float` según cómo se escribiera el archivo. Con un `str()` directo, el
        código DANE 11001 se convertía en la cadena **'11001.0'**.

        Eso no es un detalle de presentación: el código de municipio es una identidad. Con el
        decimal pegado no cruza con '11001' en la restricción única de la tabla, ni en el
        filtro por municipio con el que RF-08 recupera casos contabilizados, ni contra ningún
        catálogo. El dato entra, se ve casi bien y no sirve para nada.

        Solo se normalizan los números enteros. Un decimal real —que aquí no debería
        aparecer, pero podría en un campo mal llenado— conserva su parte decimal en vez de
        perderla en silencio.
        """
        if index is None or index >= len(row) or row[index] is None:
            return None
        valor = row[index]
        if isinstance(valor, float) and valor.is_integer():
            valor = int(valor)
        text = str(valor).strip()
        return text or None

    @staticmethod
    def _num(
        row, index: Optional[int], line: int, field: str, *, required: bool
    ) -> Optional[float]:
        if (
            index is None
            or index >= len(row)
            or row[index] is None
            or str(row[index]).strip() == ""
        ):
            if required:
                raise ValidationException(f"Fila {line}: la columna '{field}' es obligatoria")
            return None
        raw = str(row[index]).strip().replace("%", "").replace(",", ".")
        try:
            return float(raw)
        except ValueError as exc:
            raise ValidationException(
                f"Fila {line}: '{field}' debe ser numérico, se recibió '{row[index]}'"
            ) from exc
